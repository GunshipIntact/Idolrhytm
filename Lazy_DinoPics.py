import openpyxl
import os
from openpyxl_image_loader import SheetImageLoader

print("Starting image extraction...")

# Load Excel file - UPDATE THIS PATH!
wb = openpyxl.load_workbook(r'C:\Users\phill\Documents\Idolrhytm\Playing Cards.xlsx') 
sheet = wb.active

# Create folder for images
os.makedirs('card_images', exist_ok=True)

# Extract images
image_loader = SheetImageLoader(sheet)

images_saved = 0
for row in range(1, sheet.max_row + 1):
    try:
        # Assuming: Column A = images, Column B = descriptions
        description = sheet.cell(row=row, column=2).value
        
        if description and image_loader.image_in(f'A{row}'):
            image = image_loader.get(f'A{row}')
            # Clean filename (remove invalid characters)
            clean_name = "".join(c for c in str(description) if c.isalnum())
            image.save(f'card_images/{clean_name}.png')
            print(f'Saved: {clean_name}.png')
            images_saved += 1
    except Exception as e:
        print(f"Error in row {row}: {e}")

print(f"Done! Saved {images_saved} images in 'card_images' folder")